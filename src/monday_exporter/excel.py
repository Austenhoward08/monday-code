"""Excel export utilities for Monday.com boards."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Optional, Sequence, Tuple

from dateutil import parser as date_parser
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Board, BoardColumn, ColumnValue, Item

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
DATE_NUMBER_FORMAT = "yyyy-mm-dd hh:mm"
ALTERNATING_FILL = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(border_style="thin", color="D9D9D9"),
    right=Side(border_style="thin", color="D9D9D9"),
    top=Side(border_style="thin", color="D9D9D9"),
    bottom=Side(border_style="thin", color="D9D9D9"),
)


class ExcelExporter:
    """Convert Monday.com board data into a formatted Excel workbook."""

    def build_workbook(self, board: Board) -> Workbook:
        workbook = Workbook()
        items_sheet = workbook.active
        items_sheet.title = _sanitize_sheet_title(board.name) or "Board"

        summary_sheet = workbook.create_sheet("Summary")

        self._populate_items_sheet(items_sheet, board)
        self._populate_summary_sheet(summary_sheet, board)

        workbook.active = items_sheet
        return workbook

    def export(self, board: Board, destination: Path) -> Path:
        """
        Write the given board to an Excel file.

        Args:
            board: Monday.com board data.
            destination: Path to the output XLSX file.
        """
        destination = destination.expanduser().resolve()
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = self.build_workbook(board)
        workbook.save(destination)
        return destination

    # ----------------------------------------------------------------------- #
    # Sheet population helpers
    # ----------------------------------------------------------------------- #

    def _populate_items_sheet(self, sheet, board: Board) -> None:
        columns = board.ordered_columns()

        headers = [
            "Item ID",
            "Item Name",
            "Group",
            "Group Color",
            "Creator",
            "Created At",
            "Updated At",
        ] + [column.title for column in columns]

        sheet.append(headers)
        sheet.freeze_panes = "A2"

        for idx, cell in enumerate(sheet[1], start=1):
            _apply_header_style(cell)
            column_letter = get_column_letter(idx)
            sheet.column_dimensions[column_letter].width = max(len(cell.value or ""), 12)

        for item in board.items:
            row_values, column_formats = self._build_item_row(item, columns)
            sheet.append(row_values)

            row_index = sheet.max_row
            row_cells = next(sheet.iter_rows(min_row=row_index, max_row=row_index))
            is_alternate = row_index % 2 == 0

            for idx, cell in enumerate(row_cells, start=0):
                cell.border = THIN_BORDER
                cell.alignment = TEXT_ALIGNMENT

                if column_formats.get(idx) == "date":
                    cell.number_format = DATE_NUMBER_FORMAT

                if is_alternate:
                    cell.fill = ALTERNATING_FILL

                display_length = _calculate_display_width(cell.value)
                column_letter = get_column_letter(idx + 1)
                previous_width = sheet.column_dimensions[column_letter].width or 10
                if display_length > previous_width:
                    sheet.column_dimensions[column_letter].width = min(display_length, 48)

        sheet.auto_filter.ref = sheet.dimensions

    def _populate_summary_sheet(self, sheet, board: Board) -> None:
        sheet.append(["Property", "Value"])
        sheet.append(["Board Name", board.name])
        sheet.append(["Board ID", board.id])
        sheet.append(["Item Count", len(board.items)])
        sheet.append(["Group Count", len(board.groups)])
        sheet.append(["Column Count", len(board.columns)])
        if board.description:
            sheet.append(["Description", board.description])
        sheet.append(["Exported At", datetime.now(timezone.utc)])

        for cell in sheet[1]:
            _apply_header_style(cell)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            key_cell, value_cell = row
            key_cell.border = THIN_BORDER
            value_cell.border = THIN_BORDER
            value_cell.alignment = TEXT_ALIGNMENT

            if isinstance(value_cell.value, datetime):
                value_cell.number_format = DATE_NUMBER_FORMAT

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 60
        sheet.auto_filter.ref = sheet.dimensions

    # ----------------------------------------------------------------------- #
    # Row / cell helpers
    # ----------------------------------------------------------------------- #

    def _build_item_row(
        self,
        item: Item,
        columns: Sequence[BoardColumn],
    ) -> Tuple[List[Any], dict[int, str]]:
        column_formats: dict[int, str] = {}

        row: List[Any] = [
            item.id,
            item.name,
            item.group.title if item.group else "",
            item.group.color if item.group and item.group.color else "",
            item.creator.name if item.creator and item.creator.name else "",
            item.created_at or "",
            item.updated_at or "",
        ]

        if isinstance(row[5], datetime):
            column_formats[5] = "date"
        if isinstance(row[6], datetime):
            column_formats[6] = "date"

        for column in columns:
            column_value = item.column_value_by_id(column.id)
            value, format_tag = _render_column_value(column, column_value)
            if format_tag:
                column_formats[len(row)] = format_tag
            row.append(value)

        return row, column_formats


def _render_column_value(
    column: BoardColumn,
    column_value: Optional[ColumnValue],
) -> Tuple[Any, Optional[str]]:
    if column_value is None:
        return "", None

    if column.type == "date":
        parsed_date = _parse_date(column_value)
        if parsed_date:
            return parsed_date, "date"
        return column_value.text, None

    if column.type in {"numbers", "numeric"}:
        try:
            # Monday stores numeric values as strings, sometimes with commas.
            return float(column_value.text.replace(",", "")), None
        except (ValueError, AttributeError):
            return column_value.text, None

    if column.type == "checkbox":
        value = column_value.parsed_value
        if isinstance(value, dict):
            value = value.get("checked")
        if isinstance(value, bool):
            return value, None
        return column_value.text or "", None

    if column.type == "people":
        return column_value.text.replace("\n", ", "), None

    return column_value.text, None


def _parse_date(column_value: ColumnValue) -> Optional[datetime]:
    value = column_value.parsed_value

    if isinstance(value, dict):
        date_part = value.get("date")
        if date_part:
            try:
                return date_parser.isoparse(date_part)
            except (ValueError, TypeError):
                return None

    if column_value.text:
        try:
            return date_parser.parse(column_value.text)
        except (ValueError, TypeError):
            return None
    return None


def _calculate_display_width(value: Any) -> float:
    if value is None:
        return 0
    if isinstance(value, datetime):
        return 22
    text = str(value)
    lines = text.splitlines()
    max_length = max(len(line) for line in lines) if lines else 0
    return max_length + 2


def _sanitize_sheet_title(title: str) -> str:
    title = re.sub(r"[\\/*?:\\[\\]]", "_", title)
    return title[:31]


def _apply_header_style(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


__all__ = ["ExcelExporter"]
